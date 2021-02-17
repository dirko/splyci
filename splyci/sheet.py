"""
Extract constraints from a spreadsheet file.
"""

import openpyxl
import openpyxl.utils
import openpyxl.formula
import pandas


def cells_to_range(cells):
    # Note: flips position i - rows, j - cols
    cells = [(j, i) for i, j in cells]
    df = pandas.DataFrame(cells, columns=['i', 'j'])
    rowsi = sorted(df['i'].unique())
    rows = []
    rowsa = []
    for i in rowsi:
        rowsa.append(i)
        e = list(df[df.i == i].j)
        row = [e[0]]
        for _, el in enumerate(e[1:]):
            if el == row[-1] + 1:
                row.append(el)
            else:
                rows.append(row)
                row = [el]
                rowsa.append(i)
        rows.append(row)

    blocksi = []
    blocks = []
    cblocksi = [rowsa[0]]
    cblocks = rows[0]
    for i, row in zip(rowsa[1:], rows[1:]):
        if i == cblocksi[-1] + 1 and row == cblocks:
            cblocksi.append(i)
        else:
            blocksi.append(cblocksi)
            blocks.append(cblocks)
            cblocksi = [i]
            cblocks = row
    blocksi.append(cblocksi)
    blocks.append(cblocks)

    r = []
    for y, x in zip(blocksi, blocks):
        ys = min(y)
        ye = max(y)
        xs = openpyxl.utils.cell.get_column_letter(min(x) )
        xe = openpyxl.utils.cell.get_column_letter(max(x) )
        if len(x) == len(y) == 1:
            r.append(f'{xs}{ys}')
        else:
            r.append(f'{xs}{ys}:{xe}{ye}')
    return ','.join(r)





def get_indices(sheet, sheet_nr):
    edge_map = {}

    for row in sheet:
        for cell in row:
            edge_map[(cell.column, cell.row)] = (
                # For now ignore this until we have something that works better
                1 * 0,#(cell.value is None and (cell.border.left.style is not None or cell.border.right.style is not None)),
                1 * 0#(cell.value is None and (cell.border.top.style is not None or cell.border.bottom.style is not None)),
            )

    for row in sheet:
        for cell in row:
            if not cell.comment:
                continue
            anns = [tuple(ann.split(' ')) for ann in cell.comment.text.split('\n')]
            for ann in anns:
                if ann[0] == 'row_cut':
                    for (i, j) in edge_map.keys():
                        if j == cell.row and i >= cell.column and i - cell.column <= int(ann[1]):
                            eme = edge_map[i, j]
                            edge_map[i, j] = (1, eme[1])
                if ann[0] == 'col_cut':
                    for (i, j) in edge_map.keys():
                        if i == cell.column and j >= cell.row and j - cell.row <= int(ann[1]):
                            eme = edge_map[i, j]
                            edge_map[i, j] = (eme[0], 1)

    cmap = {}
    for row in sheet:
        for cell in row:
            i, j = cell.column, cell.row
            pia, _ = cmap.get((i, j - 1), (0, None))
            _, pjl = cmap.get((i - 1, j), (None, 0))
            di, dj = edge_map[(i, j)]
            cmap[i, j] = (pia + di, pjl + dj)

    index_map = {}
    for row in sheet:
        for cell in row:
            index_map[(cell.column, cell.row)] = (f'i_{sheet_nr}_{cmap[cell.column, cell.row][0]:02}_{cell.column:03}',
                                                  f'j_{sheet_nr}_{cmap[cell.column, cell.row][1]:02}_{cell.row:03}' )
    return index_map


class Formula:
    def __init__(self, v, original_i, original_j, index_map):
        self.v = v
        # First get r1c1 format
        tokenizer = openpyxl.formula.Tokenizer(self.v)
        for tok in tokenizer.items:
            if tok.subtype == 'RANGE':
                _, (x1, y1, x2, y2) = openpyxl.utils.range_to_tuple('dummy!' + tok.value)
                tok.value = f'R[{y1 - original_j}]C[{x1 - original_i}]:R[{y2 - original_j}]C[{x2 - original_i}]'
        r1c1 = tokenizer.render()

        # Now replace range with named range template {Range1}
        tokenizer = openpyxl.formula.Tokenizer(self.v)
        ranges = {}
        for tok in tokenizer.items:
            if tok.subtype == 'RANGE':
                rows = openpyxl.utils.rows_from_range(tok.value)
                cells = [openpyxl.utils.cell.coordinate_from_string(cell)
                         for row in rows for cell in list(row)]
                cells = [(openpyxl.utils.cell.column_index_from_string(cell[0]), cell[1]) for cell in cells]
                range_name = f'range_{len(ranges)}'
                tok.value = f'{{{range_name}}}'
                ranges[range_name] = [index_map[cell] for cell in cells]
        range_template = tokenizer.render()

        self.r1c1 = r1c1
        self.range_template = range_template
        self.ranges = ranges
        # Build up a cell -> set(range_name) map to quickly find the range name associated with a dependent cell
        self.dependency_backmap = {}
        for range_name, range_cells in ranges.items():
            for range_cell in range_cells:
                d = self.dependency_backmap.get(range_cell, set())
                d = d.union(range_name)
                self.dependency_backmap[range_cell] = d
        pass


def extract_cell_annotation(i, j, text):
    ans = []
    for line in text.split('\n'):
        # ans.append(tuple(line.split(' ')))
        if line.startswith('row_id'):
            ans.append(('row_id', j, line.split(' ')[1]))
        if line.startswith('col_id'):
            ans.append(('col_id', i, line.split(' ')[1]))
    return ans


def sheet_from_file(filein, sheetnr, sheet_counter):
    book = openpyxl.load_workbook(filein)
    sheet = book.worksheets[sheetnr]
    index_map = get_indices(sheet, sheet_counter)
    cells = []
    annotations = []
    for r in sheet:
        for cell in r:
            i, j = index_map[cell.column, cell.row]
            formula = None
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula = Formula(cell.value, cell.column, cell.row, index_map)
                if not formula.ranges:  # If no arguments, then handle as normal cell
                    formula = None
            annotation = extract_cell_annotation(i, j, cell.comment.text if cell.comment else '')  # Not cuts - that is already used when extracting indices
            annotations.extend(annotation)
            cells.append(Cell(cell.value, i, j, cell.column, cell.row, extract_types(cell, i, j), formula))
    return Sheet(index_map, cells, annotations)


def extract_types(cell, i, j):
    types = []
    if cell.font.i:
        types.append('italics')
    if cell.font.b:
        types.append('bold')

    theme = str(cell.fill.fgColor.theme)
    if 'Values' in theme:
        theme = 'none'
    types.append(f'simple_type(theme_{theme})')
    types.append(f'simple_type(color_{theme}_{cell.fill.fgColor.tint})'.replace('.', '_').replace('-', '_'))
    types.append(f'simple_type(cell_id_{i}_{j})')
    types.append(f'simple_type(data_type_{cell.data_type})')
    types.append(f'comp_type(column_data_type_{cell.data_type}, {i})')
    types.append(f'comp_type(row_data_type_{cell.data_type}, {j})')
    return set(types)


class Cell:
    def __init__(self, v, i, j, original_i, original_j, types=None, formula=None):
        self.v = v
        self.i = i
        self.j = j
        self.original_i = original_i
        self.original_j = original_j
        self.types = set() if types is None else types
        self.formula = formula
        self.arg_from = []

    def __repr__(self):
        return f"cell({self.i}, {self.j}, '{self.v}')"


class Sheet:
    def __init__(self, index_map, cells, annotations):
        self.cells = cells
        self.index_map = index_map
        self.annotations = annotations
        self._cell_map_original = {(c.original_i, c.original_j): c for c in cells}
        self._cell_map = {(c.i, c.j): c for c in cells}
        arg_from = self._arg_from(self._cell_map)
        for loc, froms in arg_from.items():
            self._cell_map[loc].arg_from = {self._cell_map[fr] for fr in froms}

    def __getitem__(self, item):
        if item in self._cell_map:
            return self._cell_map[item]
        if item in self._cell_map_original:
            return self._cell_map_original[item]
        return None

    def _arg_from(self, cells):
        arg_from = {}
        for (i, j), cell in cells.items():
            if cell and cell.formula:
                for range_name in cell.formula.ranges.keys():
                    for rni, rnj in cell.formula.ranges[range_name]:
                        froms = arg_from.get((rni, rnj), set())
                        froms.add((i, j))
                        arg_from[rni, rnj] = froms
        return arg_from


