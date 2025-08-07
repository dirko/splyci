import json
import collections
from dataclasses import dataclass
import datetime

import matplotlib.pyplot as plt
import matplotlib
import pandas
import numpy as np
from openpyxl.styles.borders import Border, Side

from splyci.csp import create_blocks, csp
from splyci.sheet import sheet_from_file, cells_to_range
from splyci.match import match
from splyci.block import _match_lines, _match_sets, split_lines, split_blocks, Block
from splyci.formula import generalise, FormulaBlock, FormulaBlockHorizontal, FormulaBlockVertical


def dependent_intersection(original_block, dblocks, di, dj, assignment):
    cells = []
    print('intersection', original_block, dblocks, di, dj)
    for dblock in dblocks:
        if isinstance(original_block, FormulaBlockVertical):
            for i in range(assignment[dblock.x1], assignment[dblock.x2] + 1):
                cells.append((i, dj + assignment[dblock.y1]))
        if isinstance(original_block, FormulaBlockHorizontal):
            for j in range(assignment[dblock.y1], assignment[dblock.y2] + 1):
                cells.append((di + assignment[dblock.x1], j))
    return cells


def fill_blocks(blocks, output_blocks, assignment):
    min_x = min(assignment[block.x1] for block in output_blocks) + 0
    min_y = min(assignment[block.y1] for block in output_blocks) + 0
    max_width = max(assignment[block.x2] for block in output_blocks) + 2 - min_x
    max_height = max(assignment[block.y2] for block in output_blocks) + 2 - min_y
    print('max_width, max_height', max_width, max_height, min_x, min_y)
    x = {c: [None for _ in range(max_height)] for c in range(max_width)}

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    for nblock, block in enumerate(output_blocks):
        (ids, x1, y1, x2, y2, width, height, original_blocks, dependent_blocks) = \
            block.id, block.x1, block.y1, block.x2, block.y2, block.width, block.height, block.original, block.dependencies
        print_blocks = [repr(pblock) for pblock in original_blocks]
        for assignment_k, assignment_v in assignment.items():
            print_blocks = [pblock.replace(assignment_k, f'{assignment_k}={assignment_v}') for pblock in print_blocks]
        print(nblock, '/', len(output_blocks), '  ', (width, height), print_blocks)
        # TODO how to disambiguoate?
        original_block = original_blocks[0]
        if isinstance(original_block, Block):
            # Fill in cells
            for cell in original_block.cells:
                di, dj = cell.original_i - original_block.x1o, cell.original_j - original_block.y1o
                # Note: other way around (row, column)! None - not for dicts
                x[assignment[x1] + di - min_x][assignment[y1] + dj - min_y] = cell.v
                new_cell = ws.cell(row=assignment[y1] + dj - min_y + 1, column=assignment[x1] + di - min_x + 1, value=cell.v)
                new_cell.border = _new_border(di, dj, width, height)
        else:  # isinstance(original_block, FormulaBlockVertical):
            # Fill in formulas
            for i in range(assignment[x1], assignment[x2] + 1):
                for j in range(assignment[y1], assignment[y2] + 1):
                    args = {}
                    di = i - assignment[x1]
                    dj = j - assignment[y1]
                    for range_name, v in original_block.dependencies.items():
                        print()
                        print('Range n', range_name, v, dependent_blocks)
                        # TODO think below case cannot happen?
                        # Note: it happens when the argument isn't used for placing
                        if isinstance(next(iter(v), None), Block):
                            dblocks = v
                            print('not placing', range_name)
                            icells = dependent_intersection(original_block, dblocks, di, dj, assignment)
                            drange = ([(ii - min_x+1, ij - min_y+1) for (ii, ij) in icells])
                            sorted_drange = sorted(list(set(drange)))
                            print('block drange', drange)
                            range_string = cells_to_range(sorted_drange)
                            print('block range_string', range_string)
                            args[range_name] = range_string
                        if isinstance(next(iter(v), None), str):
                            drange = []
                            if not range_name in dependent_blocks:
                                print('something went wrong with formulae - skipping this block')
                                break
                            dblocks = dependent_blocks[range_name]
                            print('rogo', original_block, range_name)
                            icells = dependent_intersection(original_block, dblocks, di, dj, assignment)
                            drange = ([(ii - min_x+1, ij - min_y+1) for (ii, ij) in icells])
                            print('drange', drange)
                            sorted_drange = sorted(list(set(drange)))
                            print('sdr', v)
                            print('sdr', sorted_drange)
                            range_string = cells_to_range(sorted_drange)
                            print('range_string', range_string)
                            args[range_name] = range_string
                        # Else the block is probably empty?

                    print('adding', args)
                    try:
                        new_value = original_block.template.format(**args)
                    except KeyError:
                        print('something went wrong with formulae - skipping this block')
                        new_value = ''

                    x[i - min_x][j - min_y] = new_value
                    new_cell = ws.cell(row=j - min_y + 1, column=i - min_x + 1, value=new_value)
                    new_cell.border = _new_border(di, dj, width, height)

    print('converting')
    df = pandas.DataFrame(x)
    return wb, df


def _new_border(di, dj, w, h):
    left = Side(border_style='thick') if di == 0 else None
    right = Side(border_style='thick') if di == w - 1 else None
    top = Side(border_style='thick') if dj == 0 else None
    bottom = Side(border_style='thick') if dj == h - 1 else None
    border = Border(left=left, right=right, top=top, bottom=bottom)
    return border


def get_index_locations(sheets):
    locations = {}
    for sheet in sheets:
        for (i, j), (ni, nj) in sheet.index_map.items():
            locations[ni] = i
            locations[nj] = j
    return locations

@dataclass
class IntegrationReport:
    num_sheet_blocks: (int, int)
    num_sheet_formula_blocks: (int, int)
    num_generalised_formula_blocks: (int, int)
    num_output_blocks: int
    num_used_blocks: int
    num_used_positional_constraints: int
    num_total_positional_constraints: int


def extract(
        filesin,
        fileout=None,
        goal=False,
        cut_use_annotations=True,
        cut_use_border_style=False,
        match_use_annotations=True,
        csp_time_limit=datetime.timedelta(seconds=60),
        csp_optimisation_level=0,
):
    sheets = [
        sheet_from_file(filein, sheetnr, sheet_counter, use_cut_annotations=cut_use_annotations, use_border_style=cut_use_border_style)
        for sheet_counter, (filein, sheetnr) in enumerate(filesin)
    ]
    index_locations = get_index_locations(sheets)
    match_tuples = match(sheets, use_match_annotations=match_use_annotations)
    match_sets = _match_sets(match_tuples)

    lines = [line for sheet in sheets for line in split_lines(sheet, match_sets, index_locations)]
    lines.extend(_match_lines(lines, match_sets, index_locations))
    sheet_blocks = [split_blocks(sheet, lines) for sheet in sheets]

    try:
        generalised_sheet_blocks = [generalise(blocks) for blocks in sheet_blocks]
        blocks = [block for blocks in generalised_sheet_blocks for block in blocks]
        output_blocks, prolog_file_name = create_blocks(blocks, match_tuples)
        ProjectedBlock = collections.namedtuple('ProjectedBlock', 'x1 y1 x2 y2 width height')
        pblocks = list(set(ProjectedBlock(block.x1, block.y1, block.x2, block.y2, block.width, block.height) for block in output_blocks))
        assignment, used_constraints = csp(pblocks, sheets, match_tuples, goal=goal, time_limit=csp_time_limit, optimisation_level=csp_optimisation_level)
        if assignment:
            wb, df = fill_blocks(blocks, output_blocks, assignment)
            pandas.options.display.width = 0
            print(df)
        integration_report = IntegrationReport(
            num_sheet_blocks=[len(sb) for sb in sheet_blocks],
            num_sheet_formula_blocks=[len([b for b in sb if b.type == 'formula']) for sb in sheet_blocks],
            num_generalised_formula_blocks=[len([b for b in sb if isinstance(b, FormulaBlock)]) for sb in generalised_sheet_blocks],
            num_output_blocks=len(output_blocks),
            num_used_blocks=len([b for b, used in used_constraints['blocks'].items() if used]),
            num_used_positional_constraints=sum(len([w for w, used in used_constraints[positional].items() if used]) for positional in ['left', 'above', 'left_equal', 'above_equal']),
            num_total_positional_constraints=sum(len([w for w, used in used_constraints[positional].items()]) for positional in ['left', 'above', 'left_equal', 'above_equal'])
        )
        save_report(integration_report, fileout + '/report.json')
        print('done')
    except Exception as error:
        print('Error')
        print(error)
        raise error
    finally:
        #print(df)
        if fileout:
            for i, blocks in enumerate(sheet_blocks):
                draw_blocks(blocks, fileout + f'/{i}.svg')
    if fileout:
        filen = fileout + '/output.xlsx'
        #df.to_excel(filen, header=False, index=False)
        wb.save(filen)
        copy_prolog_file(prolog_file_name, fileout)
    return df, integration_report


def save_report(report, filename):
    data = {
        'num_sheet_blocks': report.num_sheet_blocks,
        'num_sheet_formula_blocks': report.num_sheet_formula_blocks,
        'num_generalised_formula_blocks': report.num_generalised_formula_blocks,
        'num_output_blocks': report.num_output_blocks,
        'num_used_blocks': report.num_used_blocks,
        'num_used_positional_constraints': report.num_used_positional_constraints,
        'num_total_positional_constraints': report.num_total_positional_constraints,
    }
    with open(filename, 'w') as f:
        json.dump(data, f, indent=4, sort_keys=True, default=str)


def copy_prolog_file(filename, dir):
    with open(filename, 'r') as fi:
        with open(dir + '/blocks.pl', 'w') as fo:
            fo.write(fi.read())


def draw_blocks(blocks, filename):
    fig2 = plt.figure(figsize=(9, 8))
    ax2 = fig2.add_subplot(111, aspect='equal')
    min_x = min(b.x1o for b in blocks)
    min_y = min(b.y1o for b in blocks)
    max_width = max(b.x1o + b.width for b in blocks)
    max_height = max(b.y1o + b.height for b in blocks)

    for block in blocks:
        ax2.add_patch(
            matplotlib.patches.Rectangle(
                (block.x1o * 1.0 + np.random.random() * 0.2,
                 block.y1o * 1.0 + np.random.random() * 0.2),
                block.width * 0.96,
                block.height * 0.96,
                fill=False  # remove background
            ))

    plt.xlim(0, max_width + 1)
    plt.ylim(max_height + 1, 0)
    plt.savefig(filename)
