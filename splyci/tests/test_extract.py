from splyci.sheet import cells_to_range, Cell, Sheet, get_indices, Formula, extract_types
from splyci.match import similarity, match
from splyci.integration import sheet_from_file, get_index_locations, extract
from splyci.block import split_blocks
from splyci.csp import _parse_indices
from splyci.formula import FormulaBlockHorizontal, FormulaBlockVertical, generalise
from splyci.csp import create_blocks, csp
from unittest import TestCase
import openpyxl


class TestCellsToRange(TestCase):
    def test_cells_to_range(self):
        cells = [(1, 2), (1, 3), (1, 4)]
        r = cells_to_range(cells)
        self.assertEqual(r, 'B1:D1')

    def test_cells_to_range_v(self):
        cells = [(1, 2), (2, 2), (3, 2)]
        r = cells_to_range(cells)
        self.assertEqual(r, 'B1:B3')

    def test_cells_to_range_disj(self):
        cells = [(1, 2), (1, 3), (1, 5)]
        r = cells_to_range(cells)
        self.assertEqual(r, 'B1:C1,E1')

    def test_cells_to_range_block(self):
        cells = [(2, 5), (2, 6), (3, 5), (3, 6)]
        r = cells_to_range(cells)
        self.assertEqual(r, 'E2:F3')

    def test_cells_to_range_block_extra_h(self):
        #            E       F       E       F       G
        cells = [(2, 5), (2, 6), (3, 5), (3, 6), (2, 7)]
        r = cells_to_range(cells)
        self.assertEqual(r, 'E2:G2,E3:F3')
        #   A B C D E F G
        # 1
        # 2         X X X
        # 3         X X
        # 4

    def test_cells_to_range_block_extra_v(self):
        cells = [(2, 5), (2, 6), (3, 5), (3, 6), (4, 6)]
        r = cells_to_range(cells)
        self.assertEqual(r, 'E2:F3,F4')

    def test_cells_to_range_weird(self):
        cells = [(3, 4), (4, 4), (5, 4)]
        r = cells_to_range(cells)
        self.assertEqual(r, 'C2:E2')

    def test_cells_to_range_skip(self):
        cells = [(3, 6), (4, 6), (5, 6), (7, 6)]
        r = cells_to_range(cells)
        self.assertEqual(r, 'C6:E6,G6')


class TestMatch(TestCase):
    def test_similarity(self):
        s1 = [
            Cell('a', None, None, None, None),
            Cell('b', None, None, None, None)
        ]
        s2 = [
            Cell('a', None, None, None, None),
            Cell('c', None, None, None, None)
        ]
        s3 = [
            Cell('c', None, None, None, None),
            Cell('c', None, None, None, None)
        ]

        self.assertEqual(similarity(s1, s2), 1.0)
        self.assertEqual(similarity(s1, s3), 0.0)

    def test_match(self):
        # a b
        # c d
        #
        # a e
        # e e
        im1 = {
            (1, 1): ('c1', 'r1'),
            (2, 1): ('c2', 'r1'),
            (1, 2): ('c1', 'r2'),
            (2, 2): ('c2', 'r2'),
        }
        im2 = {
            (1, 1): ('C1', 'R1'),
            (2, 1): ('C2', 'R1'),
            (1, 2): ('C1', 'R2'),
            (2, 2): ('C2', 'R2'),
        }
        cells1 = [
            Cell('a', 'c1', 'r1', 1, 1),
            Cell('b', 'c2', 'r1', 2, 1),
            Cell('c', 'c1', 'r2', 1, 2),
            Cell('d', 'c2', 'r2', 2, 2)
        ]
        cells2 = [
            Cell('a', 'C1', 'R1', 1, 1),
            Cell('e', 'C2', 'R1', 2, 1),
            Cell('e', 'C1', 'R2', 1, 2),
            Cell('e', 'C2', 'R2', 2, 2)
        ]
        s1 = Sheet(im1, cells1)
        s2 = Sheet(im2, cells2)
        matches = match([s1, s2])
        self.assertEqual(matches, [('c1', 'C1'), ('r1', 'R1')])

    def test_match_actual(self):
        sheet1 = sheet_from_file('data/schools.xlsx', 0, 0)
        sheet2 = sheet_from_file('data/schools.xlsx', 1, 1)
        matches = match([sheet1, sheet2])
        print(matches)
        expected = sorted([('i_1_00_003', 'i_0_00_004'),
                           ('i_1_00_002', 'i_0_00_002'),
                           ('i_1_00_005', 'i_0_00_006'),
                           ('i_1_00_004', 'i_0_00_005'),
                           ('j_1_00_005', 'j_0_00_002')])
        self.assertEqual(sorted(matches), expected)


class TestExtract(TestCase):
    def test_get_indices(self):
        book = openpyxl.load_workbook('data/schools.xlsx')
        sheet = book.worksheets[0]
        index_map = get_indices(sheet, 0)
        expected = {
            (1, 1): ('i_0_00_001', 'j_0_00_001'),
            (1, 2): ('i_0_00_001', 'j_0_00_002'),
            (1, 3): ('i_0_00_001', 'j_0_00_003'),
            (1, 4): ('i_0_00_001', 'j_0_00_004'),
            (1, 5): ('i_0_00_001', 'j_0_00_005'),
            (1, 6): ('i_0_00_001', 'j_0_00_006'),
            (1, 7): ('i_0_00_001', 'j_0_00_007'),
            (1, 8): ('i_0_00_001', 'j_0_00_008'),
            (1, 9): ('i_0_00_001', 'j_0_00_009'),
            (1, 10): ('i_0_00_001', 'j_0_00_010'),
            (1, 11): ('i_0_00_001', 'j_0_00_011'),
            (1, 12): ('i_0_00_001', 'j_0_00_012'),
            (1, 13): ('i_0_00_001', 'j_0_00_013'),
            (1, 14): ('i_0_00_001', 'j_0_00_014'),
            (1, 15): ('i_0_00_001', 'j_0_00_015'),
            (2, 1): ('i_0_00_002', 'j_0_00_001'),
            (2, 2): ('i_0_00_002', 'j_0_00_002'),
            (2, 3): ('i_0_00_002', 'j_0_00_003'),
            (2, 4): ('i_0_00_002', 'j_0_00_004'),
            (2, 5): ('i_0_00_002', 'j_0_00_005'),
            (2, 6): ('i_0_00_002', 'j_0_00_006'),
            (2, 7): ('i_0_00_002', 'j_0_00_007'),
            (2, 8): ('i_0_00_002', 'j_0_00_008'),
            (2, 9): ('i_0_00_002', 'j_0_00_009'),
            (2, 10): ('i_0_00_002', 'j_0_00_010'),
            (2, 11): ('i_0_00_002', 'j_0_00_011'),
            (2, 12): ('i_0_00_002', 'j_0_00_012'),
            (2, 13): ('i_0_00_002', 'j_0_00_013'),
            (2, 14): ('i_0_00_002', 'j_0_00_014'),
            (2, 15): ('i_0_00_002', 'j_0_00_015'),
            (3, 1): ('i_0_00_003', 'j_0_00_001'),
            (3, 2): ('i_0_00_003', 'j_0_00_002'),
            (3, 3): ('i_0_00_003', 'j_0_00_003'),
            (3, 4): ('i_0_00_003', 'j_0_00_004'),
            (3, 5): ('i_0_00_003', 'j_0_00_005'),
            (3, 6): ('i_0_00_003', 'j_0_00_006'),
            (3, 7): ('i_0_00_003', 'j_0_00_007'),
            (3, 8): ('i_0_00_003', 'j_0_00_008'),
            (3, 9): ('i_0_00_003', 'j_0_00_009'),
            (3, 10): ('i_0_00_003', 'j_0_00_010'),
            (3, 11): ('i_0_00_003', 'j_0_00_011'),
            (3, 12): ('i_0_00_003', 'j_0_00_012'),
            (3, 13): ('i_0_00_003', 'j_0_00_013'),
            (3, 14): ('i_0_00_003', 'j_0_00_014'),
            (3, 15): ('i_0_00_003', 'j_0_00_015'),
            (4, 1): ('i_0_00_004', 'j_0_00_001'),
            (4, 2): ('i_0_00_004', 'j_0_00_002'),
            (4, 3): ('i_0_00_004', 'j_0_00_003'),
            (4, 4): ('i_0_00_004', 'j_0_00_004'),
            (4, 5): ('i_0_00_004', 'j_0_00_005'),
            (4, 6): ('i_0_00_004', 'j_0_00_006'),
            (4, 7): ('i_0_00_004', 'j_0_00_007'),
            (4, 8): ('i_0_00_004', 'j_0_00_008'),
            (4, 9): ('i_0_00_004', 'j_0_00_009'),
            (4, 10): ('i_0_00_004', 'j_0_00_010'),
            (4, 11): ('i_0_00_004', 'j_0_00_011'),
            (4, 12): ('i_0_00_004', 'j_0_00_012'),
            (4, 13): ('i_0_00_004', 'j_0_00_013'),
            (4, 14): ('i_0_00_004', 'j_0_00_014'),
            (4, 15): ('i_0_00_004', 'j_0_00_015'),
            (5, 1): ('i_0_00_005', 'j_0_00_001'),
            (5, 2): ('i_0_00_005', 'j_0_00_002'),
            (5, 3): ('i_0_00_005', 'j_0_00_003'),
            (5, 4): ('i_0_00_005', 'j_0_00_004'),
            (5, 5): ('i_0_00_005', 'j_0_00_005'),
            (5, 6): ('i_0_00_005', 'j_0_00_006'),
            (5, 7): ('i_0_00_005', 'j_0_00_007'),
            (5, 8): ('i_0_00_005', 'j_0_00_008'),
            (5, 9): ('i_0_00_005', 'j_0_00_009'),
            (5, 10): ('i_0_00_005', 'j_0_00_010'),
            (5, 11): ('i_0_00_005', 'j_0_00_011'),
            (5, 12): ('i_0_00_005', 'j_0_00_012'),
            (5, 13): ('i_0_00_005', 'j_0_00_013'),
            (5, 14): ('i_0_00_005', 'j_0_00_014'),
            (5, 15): ('i_0_00_005', 'j_0_00_015'),
            (6, 1): ('i_0_00_006', 'j_0_00_001'),
            (6, 2): ('i_0_00_006', 'j_0_00_002'),
            (6, 3): ('i_0_00_006', 'j_0_00_003'),
            (6, 4): ('i_0_00_006', 'j_0_00_004'),
            (6, 5): ('i_0_00_006', 'j_0_00_005'),
            (6, 6): ('i_0_00_006', 'j_0_00_006'),
            (6, 7): ('i_0_00_006', 'j_0_00_007'),
            (6, 8): ('i_0_00_006', 'j_0_00_008'),
            (6, 9): ('i_0_00_006', 'j_0_00_009'),
            (6, 10): ('i_0_00_006', 'j_0_00_010'),
            (6, 11): ('i_0_00_006', 'j_0_00_011'),
            (6, 12): ('i_0_00_006', 'j_0_00_012'),
            (6, 13): ('i_0_00_006', 'j_0_00_013'),
            (6, 14): ('i_0_00_006', 'j_0_00_014'),
            (6, 15): ('i_0_00_006', 'j_0_00_015'),
            (7, 1): ('i_0_01_007', 'j_0_01_001'),
            (7, 2): ('i_0_01_007', 'j_0_00_002'),
            (7, 3): ('i_0_01_007', 'j_0_00_003'),
            (7, 4): ('i_0_01_007', 'j_0_00_004'),
            (7, 5): ('i_0_01_007', 'j_0_00_005'),
            (7, 6): ('i_0_01_007', 'j_0_00_006'),
            (7, 7): ('i_0_01_007', 'j_0_00_007'),
            (7, 8): ('i_0_01_007', 'j_0_00_008'),
            (7, 9): ('i_0_01_007', 'j_0_00_009'),
            (7, 10): ('i_0_01_007', 'j_0_00_010'),
            (7, 11): ('i_0_01_007', 'j_0_00_011'),
            (7, 12): ('i_0_01_007', 'j_0_00_012'),
            (7, 13): ('i_0_01_007', 'j_0_00_013'),
            (7, 14): ('i_0_01_007', 'j_0_00_014'),
            (7, 15): ('i_0_01_007', 'j_0_00_015')
        }
        self.assertEqual(index_map, expected)

    def test_formula(self):
        im = {
            (1, 1): ('c1', 'r1'),
            (2, 1): ('c2', 'r1'),
            (1, 2): ('c1', 'r2'),
            (2, 2): ('c2', 'r2'),
            (5, 10): ('e1', 'e2')
        }
        f = Formula('=SUM(A1:B2) + E10', 2, 3, im)
        self.assertEqual(f.r1c1, '=SUM(R[-2]C[-1]:R[-1]C[0]) + R[7]C[3]:R[7]C[3]')
        self.assertEqual(f.ranges, {'range_0': [('c1', 'r1'), ('c2', 'r1'), ('c1', 'r2'), ('c2', 'r2')],
                                    'range_1': [('e1', 'e2')]})

    def test_sheet_from_file(self):
        sheet = sheet_from_file('data/schools.xlsx', 0, 4)
        self.assertEqual(len(sheet.cells), 105)
        self.assertEqual(len(sheet.index_map), 105)

    def test_extract_types(self):
        book = openpyxl.load_workbook('data/schools.xlsx')
        sheet = book.worksheets[0]
        cell = sheet.cell(3, 4)
        types = extract_types(cell)
        self.assertEqual(types, {'theme_3', 'color_3_0_3999755851924192'})


class TestBlocks(TestCase):
    def test_split_blocks(self):
        im = {
            (1, 1): ('c1', 'r1'),
            (2, 1): ('c2', 'r1'),
            (1, 2): ('c1', 'r2'),
            (2, 2): ('c2', 'r2'),
        }
        cells = [
            Cell('a', 'c1', 'r1', 1, 1),
            Cell('b', 'c2', 'r1', 2, 1),
            Cell('c', 'c1', 'r2', 1, 2),
            Cell('d', 'c2', 'r2', 2, 2)
        ]
        matches = []
        index_locations = {v: k for k, v in im.items()}
        sheet = Sheet(im, cells)
        blocks = split_blocks(sheet, matches, index_locations)
        self.assertEqual(len(blocks), 1)
        self.assertEqual(len(blocks[0].cells), 4)

    def test_split_actual(self):
        sheet = sheet_from_file('data/schools.xlsx', 0, 0)
        index_map = sheet.index_map
        matches = {}
        index_locations = {v: k for k, v in index_map.items()}
        blocks = split_blocks(sheet, matches, index_locations)
        print(blocks)
        self.assertEqual(len(blocks), 9)
        self.assertEqual(len(blocks[0].cells), 6)
        self.assertEqual(len(blocks[1].cells), 3)

    def test_split_actual_match(self):
        sheet1 = sheet_from_file('data/schools.xlsx', 0, 0)
        sheet2 = sheet_from_file('data/schools.xlsx', 1, 1)
        matches = match([sheet1, sheet2])
        index_locations = get_index_locations([sheet1, sheet2])
        blocks = split_blocks(sheet1, matches, index_locations)
        print(matches)
        print(blocks)
        self.assertEqual(len(blocks), 11)
        self.assertEqual(len(blocks[0].cells), 2)
        self.assertEqual(blocks[1].types, {'italics', 'color_3_0_7999816888943144', 'theme_3', 'bold'})
        self.assertEqual(len(blocks[1].cells), 1)

    def test_split_empty(self):
        sheet = sheet_from_file('data/schools.xlsx', 5, 0)
        index_map = sheet.index_map
        matches = {}
        index_locations = {v: k for k, v in index_map.items()}
        print(index_locations)
        blocks = split_blocks(sheet, matches, index_locations)
        print(blocks)
        self.assertEqual(len(blocks), 1)
        self.assertEqual(len(blocks[0].cells), 12)

    def test_split_heading(self):
        sheet = sheet_from_file('data/schools.xlsx', 6, 6)
        index_locations = get_index_locations([sheet])
        matches = {}
        sheet_blocks = split_blocks(sheet, matches, index_locations)
        print(sheet_blocks)
        self.assertEqual(len(sheet_blocks), 2)

    def test_split_heading_match(self):
        sheet1 = sheet_from_file('data/schools.xlsx', 6, 6)
        sheet2 = sheet_from_file('data/schools.xlsx', 7, 7)
        sheets = [sheet1, sheet2]
        index_locations = get_index_locations(sheets)
        matches = match(sheets)
        sheet_blocks = split_blocks(sheet1, matches, index_locations)
        print(sheet_blocks)
        self.assertEqual(len(sheet_blocks), 3)


class TestFormula(TestCase):
    def test_generalise(self):
        sheet = sheet_from_file('data/schools.xlsx', 0, 0)
        index_locations = get_index_locations([sheet])
        matches = {}
        blocks = split_blocks(sheet, matches, index_locations)
        gblocks = generalise(blocks)
        print(gblocks)
        self.assertEqual(len(gblocks), 8)
        self.assertIsInstance(gblocks[-1], FormulaBlockHorizontal)
        self.assertEqual(gblocks[-1].dependant_types, {'color_3_0_3999755851924192', 'color_3_0_5999938962981048'})
        self.assertIsInstance(gblocks[-2], FormulaBlockVertical)


class TestCsp(TestCase):
    def test_create_blocks(self):
        sheet = sheet_from_file('data/schools.xlsx', 0, 0)
        index_locations = get_index_locations([sheet])
        matches = {}
        blocks = split_blocks(sheet, matches, index_locations)
        gblocks = generalise(blocks)
        output_blocks = create_blocks(gblocks, matches)
        self.assertEqual(len(output_blocks), 9)

    def test_csp(self):
        sheet = sheet_from_file('data/schools.xlsx', 0, 0)
        index_locations = get_index_locations([sheet])
        matches = {}
        blocks = split_blocks(sheet, matches, index_locations)
        output_blocks = create_blocks(blocks, matches)
        assignment = csp(output_blocks, [sheet], matches)
        expected = {
            'i_0_00_001': 8,
            'i_0_00_003': 10,
            'i_0_00_004': 11,
            'i_0_00_006': 13,
            'i_0_01_007': 14,
            'j_0_00_001': 12,
            'j_0_00_002': 13,
            'j_0_00_003': 14,
            'j_0_00_005': 16,
            'j_0_00_006': 17,
            'j_0_01_001': 12}
        self.assertEqual(assignment, expected)

    def test_parse_indices(self):
        ti = ['xindex(=, kaas)', 'yindex(>, yindex(<, j12))']
        al, ae = _parse_indices(ti, [('j12', 'j12m')])
        self.assertEqual(al, [('yindex(<, j12)', 'yindex(>, yindex(<, j12))'), ('yindex(<, j12)', 'j12')])
        self.assertEqual(ae, [('xindex(=, kaas)', 'kaas'), ('j12', 'j12m')])


class TestEndToEnd(TestCase):
    def test_extract_single(self):
        df = extract([('data/schools.xlsx', 0)], fileout=None)
        self.assertEqual(df.iloc[1, 1], 'SchoolID')

    def test_extract(self):
        extract([('test_data.xlsx', 0), ('test_data.xlsx', 1)], fileout=None)

    def test_cut(self):
        df = extract([('test_data.xlsx', 6)], fileout=None)
        self.assertEqual(df.iloc[2, 0], 'a')

    def test_heading(self):
        df = extract([('test_data.xlsx', 6), ('test_data.xlsx', 7)], fileout=None)
        self.assertEqual(df.iloc[2, 0], 'a')
        self.assertEqual(df.iloc[4, 1], 4)

    def test_extract_small_one(self):
        extract([('test_data.xlsx', 3)], fileout=None)

    def test_extract_small(self):
        extract([('test_data.xlsx', 3), ('test_data.xlsx', 4)], fileout=None)

    def test_extract_lines_matches(self):
        extract([('test_data.xlsx', 8), ('test_data.xlsx', 9)], fileout=None)

    def test_side_formula(self):
        df = extract([('test_data.xlsx', 10)], fileout=None)
        self.assertEqual(df.iloc[1, 3], '=C2*10')

    def test_multi_block_dep(self):
        df = extract([('test_data.xlsx', 11)], fileout=None)
        self.assertEqual(df.iloc[1, 2], '=A2+B2')

    def test_multi_block_dep_merge(self):
        df = extract([('test_data.xlsx', 11), ('test_data.xlsx', 13)], fileout=None)
        self.assertEqual(df.iloc[3, 2], '=A4+B4')

    def test_split_range(self):
        df = extract([('test_data.xlsx', 12)], fileout=None)
        self.assertEqual(df.iloc[1, 2], '=A2+B2')

    def test_depends_multiple_formulas(self):
        df = extract([('test_data.xlsx', 15)], fileout=None)
        self.assertEqual(df.iloc[1, 3], '=C2+B2')

    def test_merge_dependent_formulas(self):
        df = extract([('test_data.xlsx', 13), ('test_data.xlsx', 14)], fileout=None)
        self.assertEqual(df.iloc[1, 3], '=C2*10')

    def test_match_best_only(self):
        df = extract([('test_data.xlsx', 18), ('test_data.xlsx', 19)], fileout=None)
        self.assertEqual(df.iloc[4, 3], 5)

    def test_match_hierarchical_headers(self):
        df = extract([('test_data.xlsx', 20), ('test_data.xlsx', 21)], fileout=None)
        self.assertEqual(df.iloc[3, 3], 7)
